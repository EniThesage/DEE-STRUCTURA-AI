from decimal import Decimal, ROUND_HALF_UP

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

AMOUNT_FORMAT = '#,##0'
TABLE_HEADERS = ['Item', 'Description', 'Qty', 'Unit', 'Rate (₦)', 'Amount (₦)']
COLUMN_WIDTHS = [8, 55, 10, 8, 14, 16]


def _money(value):
    return value.quantize(Decimal('1'), rounding=ROUND_HALF_UP)


def _label_row(ws, row, label, bold=False, italic=False, size=None, merge_from=1, merge_to=5, center=False):
    ws.merge_cells(start_row=row, start_column=merge_from, end_row=row, end_column=merge_to)
    cell = ws.cell(row=row, column=merge_from, value=label)
    cell.font = Font(bold=bold, italic=italic, size=size or 11)
    if center:
        cell.alignment = Alignment(horizontal='center')
    return cell


def _money_cell(ws, row, col, value, bold=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.number_format = AMOUNT_FORMAT
    if bold:
        cell.font = Font(bold=True)
    return cell


def build_beme_workbook(project, document, elements_data, bill_description):
    wb = Workbook()
    ws = wb.active
    ws.title = 'BOQ'

    for idx, width in enumerate(COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    row = _write_letterhead(ws, 1, document)
    row = _write_title(ws, row, project, bill_description, document)

    element_totals = []
    for entry in elements_data:
        row, total = _write_element(ws, row, entry['element'], entry['pages'])
        element_totals.append((entry['element'].element_number, entry['element'].title, total))
        row += 1

    row, summary_total = _write_summary(ws, row, element_totals)
    row += 1
    row = _write_general_summary(ws, row, summary_total, document)
    row += 1
    _write_footer(ws, row, document)

    return wb


def _write_letterhead(ws, row, document):
    _label_row(ws, row, document.letterhead_company or 'DEE STRUCTURA AI', bold=True, size=14, merge_to=6, center=True)
    row += 1
    if document.letterhead_address:
        _label_row(ws, row, document.letterhead_address, merge_to=6, center=True)
        row += 1
    contact_bits = [bit for bit in (document.letterhead_phone, document.letterhead_email) if bit]
    if contact_bits:
        _label_row(ws, row, ' · '.join(contact_bits), merge_to=6, center=True)
        row += 1
    return row + 1


def _write_title(ws, row, project, bill_description, document):
    _label_row(ws, row, project.name.upper(), bold=True, size=12, merge_to=6, center=True)
    row += 1
    _label_row(ws, row, f'BILL NO. 1: {bill_description}', merge_to=6, center=True)
    row += 1
    if document.reference_number:
        _label_row(ws, row, f'Ref: {document.reference_number}', merge_to=6, center=True)
        row += 1
    return row + 1


def _write_element(ws, row, element, pages):
    _label_row(ws, row, f'ELEMENT NO. {element.element_number} — {element.title.upper()}', bold=True, size=12, merge_to=6)
    row += 1

    page_subtotals = []
    for page in pages:
        for col, heading in enumerate(TABLE_HEADERS, start=1):
            ws.cell(row=row, column=col, value=heading).font = Font(bold=True)
        row += 1

        for line in page['lines']:
            if line.is_section_header:
                _label_row(ws, row, line.description, italic=True, merge_from=2, merge_to=6)
                row += 1
                continue

            ws.cell(row=row, column=1, value=line.item_label)
            ws.cell(row=row, column=2, value=line.description)
            if line.qty is not None:
                ws.cell(row=row, column=3, value=float(line.qty))
            ws.cell(row=row, column=4, value=line.unit)
            if line.rate is not None:
                _money_cell(ws, row, 5, float(line.rate))
            if line.amount is not None:
                _money_cell(ws, row, 6, float(line.amount))
            row += 1

        _label_row(ws, row, 'Carried to Collection', bold=True, merge_to=5)
        _money_cell(ws, row, 6, float(page['subtotal']), bold=True)
        page_subtotals.append((page['code'], page['subtotal']))
        row += 1

    _label_row(ws, row, 'COLLECTION', bold=True, merge_to=6)
    row += 1

    for code, subtotal in page_subtotals:
        _label_row(ws, row, f'Page {code}', merge_to=5)
        _money_cell(ws, row, 6, float(subtotal))
        row += 1

    element_total = sum((subtotal for _, subtotal in page_subtotals), Decimal('0'))
    _label_row(ws, row, 'ELEMENT TOTAL Carried to Summary', bold=True, merge_to=5)
    _money_cell(ws, row, 6, float(element_total), bold=True)
    row += 1

    return row, element_total


def _write_summary(ws, row, element_total_cells):
    _label_row(ws, row, 'SUMMARY', bold=True, size=12, merge_to=6)
    row += 1
    for col, heading in enumerate(['S/No', 'Description', '', '', '', 'Amount (₦)'], start=1):
        if heading:
            ws.cell(row=row, column=col, value=heading).font = Font(bold=True)
    row += 1

    for number, title, total in element_total_cells:
        ws.cell(row=row, column=1, value=number)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        ws.cell(row=row, column=2, value=title.upper())
        _money_cell(ws, row, 6, float(total))
        row += 1

    summary_total = sum((total for _, _, total in element_total_cells), Decimal('0'))
    _label_row(ws, row, 'TOTAL AMOUNT', bold=True, merge_to=5)
    _money_cell(ws, row, 6, float(summary_total), bold=True)
    row += 1

    return row, summary_total


def _write_general_summary(ws, row, summary_total, document):
    _label_row(ws, row, 'GENERAL SUMMARY', bold=True, size=12, merge_to=6)
    row += 1

    for col, heading in enumerate(['S/No', 'Description', '', '', '', 'Amount (₦)'], start=1):
        if heading:
            ws.cell(row=row, column=col, value=heading).font = Font(bold=True)
    row += 1

    ws.cell(row=row, column=1, value=1)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    ws.cell(row=row, column=2, value='Main Building')
    _money_cell(ws, row, 6, float(summary_total))
    row += 1

    _label_row(ws, row, 'SUB-TOTAL', bold=True, merge_to=5)
    _money_cell(ws, row, 6, float(summary_total), bold=True)
    row += 1

    preliminaries = _money(document.preliminaries)
    contingency = _money(document.contingency_amount)
    professional_fees = _money(document.professional_fees_amount)
    vat = _money(document.vat_amount)

    _label_row(ws, row, f'ADD: PRELIMINARIES ({document.preliminaries_percent}%)', merge_to=5)
    _money_cell(ws, row, 6, float(preliminaries))
    row += 1

    _label_row(ws, row, f'ADD: CONTINGENCY ({document.contingency_percent}%)', merge_to=5)
    _money_cell(ws, row, 6, float(contingency))
    row += 1

    _label_row(ws, row, f'ADD: PROFESSIONAL FEES ({document.professional_fees_percent}%)', merge_to=5)
    _money_cell(ws, row, 6, float(professional_fees))
    row += 1

    _label_row(ws, row, f'ADD: VAT ({document.vat_percent}%)', merge_to=5)
    _money_cell(ws, row, 6, float(vat))
    row += 1

    _label_row(ws, row, 'TOTAL CONTRACT SUM', bold=True, merge_to=5)
    total_contract_sum = summary_total + preliminaries + contingency + professional_fees + vat
    _money_cell(ws, row, 6, float(total_contract_sum), bold=True)
    row += 1

    return row


def _write_footer(ws, row, document):
    _label_row(ws, row, f'Prepared by: {document.prepared_by or "_______________________"}', merge_to=6)
    row += 1
    _label_row(ws, row, f'Date: {document.generated_at.strftime("%d %b %Y")}', merge_to=6)
